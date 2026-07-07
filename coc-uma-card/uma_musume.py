from __future__ import annotations

import base64
import datetime as dt
import json
import math
import os
import random
import re
import shutil
import subprocess
import tempfile
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import requests
from flask import Flask, jsonify, request, send_file
from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parent
ASSET_DIR = ROOT / "assets"
SAVE_DIR = ROOT / "achievements"
STABLE_DIR = SAVE_DIR / "stables"
HORSE_TEAM_DIR = SAVE_DIR / "horse_teams"
RACE_GROUP_DIR = SAVE_DIR / "race_groups"
RACE_DIR = SAVE_DIR / "races"
TEMPLATE_PATH = ASSET_DIR / "uma_final_template_blank.png"
MAP_PATH = ASSET_DIR / "uma_final_template_map.json"
RANK_DIR = ASSET_DIR / "rank_badges"
STAT_DIR = ASSET_DIR / "stat_badges"

FONT_CANDIDATES = [
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
    "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc",
    "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
    "/System/Library/Fonts/Hiragino Sans GB.ttc",
    "/System/Library/Fonts/STHeiti Medium.ttc",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
]

BROWN = (103, 45, 0)
WHITE = (255, 255, 255)
GRADE_COLORS = {
    "SS": (245, 168, 28),
    "S": (53, 170, 18),
    "A": (238, 102, 36),
    "B": (220, 72, 120),
    "C": (99, 190, 70),
    "D": (72, 134, 215),
    "E": (116, 116, 126),
    "F": (112, 112, 112),
    "G": (84, 84, 84),
}

RANK_THRESHOLDS = [
    ("UG", 23900),
    ("SS", 19600),
    ("S", 17500),
    ("A", 14500),
    ("B", 12500),
    ("C", 10000),
    ("D", 7000),
    ("E", 5000),
    ("F", 3000),
    ("G", 0),
]

STAT_LABELS = {
    "speed": "速度",
    "stamina": "耐力",
    "power": "力量",
    "guts": "根性",
    "wisdom": "智力",
}

TEAM_SLOTS = {
    "upper": "上等",
    "middle": "中等",
    "lower": "下等",
}

SLOT_ALIASES = {
    "upper": "upper",
    "上等": "upper",
    "上": "upper",
    "上马": "upper",
    "high": "upper",
    "middle": "middle",
    "中等": "middle",
    "中": "middle",
    "中马": "middle",
    "mid": "middle",
    "lower": "lower",
    "下等": "lower",
    "下": "lower",
    "下马": "lower",
    "low": "lower",
}

ORDER_CHAR_MAP = {
    "上": "upper",
    "中": "middle",
    "下": "lower",
}

SKILL_DEFAULTS = {
    "攀爬": 20,
    "跳跃": 20,
    "游泳": 20,
    "潜水": 1,
    "汽车驾驶": 20,
    "驾驶": 1,
    "骑术": 5,
    "导航": 10,
    "追踪": 10,
    "斗殴": 25,
    "格斗": 25,
    "近战": 25,
    "射击": 20,
    "手枪": 20,
    "步枪": 25,
    "霰弹枪": 25,
    "弓": 15,
    "投掷": 20,
    "爆破": 1,
    "炮术": 1,
    "取悦": 15,
    "恐吓": 15,
    "话术": 5,
    "说服": 10,
    "信用": 0,
    "急救": 30,
    "医学": 1,
    "精神分析": 1,
    "催眠": 1,
    "侦查": 25,
    "聆听": 20,
    "心理学": 10,
    "潜行": 20,
    "妙手": 10,
    "锁匠": 1,
    "电气": 10,
    "机械": 10,
    "电子": 1,
    "计算机": 5,
    "计算": 5,
    "乔装": 5,
    "图书馆": 20,
    "读唇": 1,
    "估价": 5,
    "会计": 5,
    "法律": 5,
    "外语": 1,
    "母语": 0,
    "克苏鲁": 0,
    "生存": 10,
    "重型": 1,
    "驯兽": 5,
    "人类学": 1,
    "历史": 5,
    "考古": 1,
    "博物": 10,
    "自然": 10,
    "神秘": 5,
    "科学": 1,
}

STAT_SKILL_GROUPS = {
    "speed": ["闪避", "跳跃", "攀爬", "汽车驾驶", "驾驶", "骑术"],
    "stamina": ["游泳", "潜水", "生存", "急救"],
    "power": ["斗殴", "格斗", "近战", "投掷", "重型"],
    "guts": ["精神分析", "克苏鲁", "急救", "医学", "生存"],
    "wisdom": ["图书馆", "侦查", "聆听", "心理学", "历史", "神秘", "科学", "计算机", "会计", "法律", "语言", "外语"],
}

APTITUDE_GROUPS = {
    "melee": ["斗殴", "格斗", "近战", "刀", "剑", "斧", "矛"],
    "ranged": ["射击", "手枪", "步枪", "霰弹枪", "弓", "投掷"],
    "city": ["侦查", "聆听", "心理学", "话术", "说服", "取悦", "恐吓", "信用"],
    "wild": ["追踪", "导航", "生存", "自然", "博物", "潜行", "攀爬"],
    "library": ["图书馆", "历史", "神秘", "考古", "人类学", "科学"],
    "data": ["计算机", "计算", "会计", "法律", "外语", "母语", "图书馆"],
    "endurance": ["急救", "医学", "生存"],
    "emergency": ["急救", "机械", "电气", "电子", "驾驶"],
    "mental": ["精神分析", "克苏鲁"],
    "luck": ["幸运"],
}


app = Flask(__name__)
for directory in (SAVE_DIR, STABLE_DIR, HORSE_TEAM_DIR, RACE_GROUP_DIR, RACE_DIR):
    directory.mkdir(exist_ok=True)

INVALID_XML_CHARS_RE = re.compile(
    "[^\u0009\u000A\u000D\u0020-\uD7FF\uE000-\uFFFD\U00010000-\U0010FFFF]"
)
BARE_AMP_RE = re.compile(r"&(?!amp;|lt;|gt;|quot;|apos;|#[0-9]+;|#x[0-9a-fA-F]+;)")
DATA_VALIDATIONS_RE = re.compile(
    r"<(?:[A-Za-z_][\w.-]*:)?dataValidations\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?dataValidations>)",
    re.DOTALL,
)
CONDITIONAL_FORMATTING_RE = re.compile(
    r"<(?:[A-Za-z_][\w.-]*:)?conditionalFormatting\b[^>]*(?:/>|>.*?</(?:[A-Za-z_][\w.-]*:)?conditionalFormatting>)",
    re.DOTALL,
)
STYLE_REFERENCE_RE = re.compile(r'(<(?:[A-Za-z_][\w.-]*:)?(?:c|row|col)\b[^>]*)\s+(?:s|style)="[^"]*"')
MINIMAL_STYLES_XML = b'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1">
    <font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/><scheme val="minor"/></font>
  </fonts>
  <fills count="2">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
  </fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
  <dxfs count="0"/>
  <tableStyles count="0" defaultTableStyle="TableStyleMedium9" defaultPivotStyle="PivotStyleLight16"/>
</styleSheet>'''


def load_field_map() -> dict[str, Any]:
    if not MAP_PATH.exists():
        raise FileNotFoundError(f"缺少坐标表: {MAP_PATH}")
    return json.loads(MAP_PATH.read_text(encoding="utf-8"))


FIELDS = load_field_map()


def get_font(size: int) -> ImageFont.FreeTypeFont:
    for path in FONT_CANDIDATES:
        if Path(path).exists():
            return ImageFont.truetype(path, size=size)
    return ImageFont.load_default(size=size)


def clean_key(value: Any) -> str:
    return re.sub(r"[^\u4e00-\u9fa5a-zA-Z]", "", str(value or "")).lower()


def parse_num(value: Any) -> float | None:
    if isinstance(value, (int, float)):
        return float(value)
    match = re.search(r"-?\d+(?:\.\d+)?", str(value or ""))
    return float(match.group()) if match else None


def get_next_value(sheet, row_idx: int, col_idx: int, max_col: int, *, expect_num=False, max_offset=5):
    for offset in range(1, max_offset + 1):
        if col_idx + offset > max_col:
            break
        raw = sheet.cell(row=row_idx, column=col_idx + offset).value
        if raw is None or str(raw).strip() == "":
            continue
        if expect_num:
            num = parse_num(raw)
            if num is not None:
                return num
            continue
        return str(raw).strip()
    return None


def get_base_value(sheet, row_idx: int, col_idx: int, max_col: int) -> int | None:
    nums = []
    for offset in range(1, 5):
        if col_idx + offset > max_col:
            break
        num = parse_num(sheet.cell(row=row_idx, column=col_idx + offset).value)
        if num is not None and 1 <= num <= 130:
            nums.append(int(num))
    below = parse_num(sheet.cell(row=row_idx + 1, column=col_idx).value)
    if below is not None and 1 <= below <= 130:
        nums.append(int(below))
    return max(nums) if nums else None


def get_skill_total(sheet, row_idx: int, col_idx: int, max_col: int) -> int | None:
    nums = []
    for offset in range(1, max_col - col_idx + 1):
        num = parse_num(sheet.cell(row=row_idx, column=col_idx + offset).value)
        if num is not None and 0 <= num <= 130:
            nums.append(int(num))
    if not nums:
        return None
    num_set = set(nums)
    for value in sorted(num_set, reverse=True):
        if 5 <= value <= 99 and (value // 2) in num_set and (value // 5) in num_set:
            return value
    for value in sorted(num_set, reverse=True):
        if 2 <= value <= 99 and (value // 2) in num_set:
            return value
    value = max(nums)
    return value if 0 <= value <= 99 else None


def find_libreoffice_binary() -> str | None:
    for name in ("soffice", "libreoffice", "soffice.exe", "soffice.com"):
        found = shutil.which(name)
        if found:
            return found
    mac_path = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    if os.path.exists(mac_path):
        return mac_path
    return None


def recalculate_or_convert_excel(file_path: str) -> str:
    suffix = Path(file_path).suffix.lower()
    binary = find_libreoffice_binary()
    if not binary:
        if suffix == ".xls":
            raise ValueError("当前环境没有 LibreOffice，暂时不能读取 xls。请换成 xlsx。")
        return file_path

    out_dir = tempfile.mkdtemp(prefix="uma_recalc_")
    profile_dir = tempfile.mkdtemp(prefix="uma_lo_profile_")
    user_profile = "-env:UserInstallation=file:///" + profile_dir.replace("\\", "/")
    try:
        subprocess.run(
            [binary, user_profile, "--headless", "--calc", "--convert-to", "xlsx", "--outdir", out_dir, file_path],
            capture_output=True,
            timeout=60,
            check=False,
        )
    except Exception:
        return file_path
    converted = Path(out_dir) / f"{Path(file_path).stem}.xlsx"
    return str(converted) if converted.exists() else file_path


def clean_xml_payload(
    data: bytes,
    *,
    strip_data_validations: bool = False,
    strip_conditional_formatting: bool = False,
    strip_cell_styles: bool = False,
) -> bytes:
    text = data.decode("utf-8", errors="replace")
    text = INVALID_XML_CHARS_RE.sub("", text)
    text = BARE_AMP_RE.sub("&amp;", text)
    if strip_data_validations:
        text = DATA_VALIDATIONS_RE.sub("", text)
    if strip_conditional_formatting:
        text = CONDITIONAL_FORMATTING_RE.sub("", text)
    if strip_cell_styles:
        while True:
            cleaned = STYLE_REFERENCE_RE.sub(r"\1", text)
            if cleaned == text:
                break
            text = cleaned
    return text.encode("utf-8")


def repair_xlsx_xml(file_path: str) -> str:
    if Path(file_path).suffix.lower() != ".xlsx":
        return file_path

    fd, repaired_path = tempfile.mkstemp(prefix="uma_repair_", suffix=".xlsx")
    os.close(fd)
    try:
        with zipfile.ZipFile(file_path, "r") as zin, zipfile.ZipFile(repaired_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                lower_name = item.filename.lower()
                if lower_name == "xl/styles.xml":
                    data = MINIMAL_STYLES_XML
                    zout.writestr(item, data)
                    continue
                if lower_name.endswith(".xml") or lower_name.endswith(".rels"):
                    is_sheet = lower_name.startswith("xl/worksheets/")
                    data = clean_xml_payload(
                        data,
                        strip_data_validations=is_sheet,
                        strip_conditional_formatting=is_sheet,
                        strip_cell_styles=is_sheet,
                    )
                zout.writestr(item, data)
        return repaired_path
    except Exception:
        if os.path.exists(repaired_path):
            os.remove(repaired_path)
        return file_path


def load_workbook_with_repair(file_path: str):
    file_path = recalculate_or_convert_excel(file_path)
    try:
        return openpyxl.load_workbook(file_path, data_only=True), None
    except Exception as first_error:
        repaired_path = repair_xlsx_xml(file_path)
        if repaired_path == file_path:
            raise first_error
        try:
            return openpyxl.load_workbook(repaired_path, data_only=True), repaired_path
        except Exception as second_error:
            if os.path.exists(repaired_path):
                os.remove(repaired_path)
            raise first_error from second_error


def normalize_skill_key(key: str) -> str:
    key = re.sub(r"[a-z]+$", "", key)
    key = key.replace("圖書館", "图书馆").replace("偵查", "侦查")
    key = key.replace("聆聽", "聆听").replace("會計", "会计")
    return key


def parse_coc_excel(file_path: str) -> dict[str, Any]:
    workbook, repaired_path = load_workbook_with_repair(file_path)
    target_sheet = None

    for sheet_name in workbook.sheetnames:
        if any(word in sheet_name for word in ("人物", "角色", "调查员", "核心")):
            target_sheet = workbook[sheet_name]
            break

    if target_sheet is None:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=20, values_only=True):
                if any(clean_key(cell) in ("力量", "str") for cell in row):
                    target_sheet = sheet
                    break
            if target_sheet is not None:
                break

    if target_sheet is None:
        target_sheet = workbook.active

    attrs = {
        "name": "调查员",
        "player": "",
        "profession": "探索者",
        "san": 0,
        "luck": 0,
        "bases": {"str": 0, "dex": 0, "pow": 0, "int": 0, "con": 0, "app": 0, "siz": 0, "edu": 0},
        "skills": {},
    }

    base_keys = {
        "str": ("力量", "str"),
        "dex": ("敏捷", "dex"),
        "pow": ("意志", "pow"),
        "int": ("智力", "灵感", "int"),
        "con": ("体质", "con"),
        "app": ("外貌", "app"),
        "siz": ("体型", "siz"),
        "edu": ("教育", "知识", "edu"),
    }
    all_skill_aliases = sorted({alias for aliases in list(STAT_SKILL_GROUPS.values()) + list(APTITUDE_GROUPS.values()) for alias in aliases})

    max_col = target_sheet.max_column
    found_name = found_player = found_profession = False

    for row_idx, row in enumerate(target_sheet.iter_rows(values_only=False), 1):
        for col_idx, cell in enumerate(row, 1):
            raw = str(cell.value or "").strip()
            if not raw:
                continue
            key = clean_key(raw)
            if not key:
                continue

            if not found_player and key in ("玩家", "玩家名", "pl", "pl名称", "player"):
                value = get_next_value(target_sheet, row_idx, col_idx, max_col)
                if value and parse_num(value) is None:
                    attrs["player"] = str(value).strip()
                    found_player = True
                continue

            if not found_name and key in ("姓名", "角色名", "调查员", "调查员姓名"):
                value = get_next_value(target_sheet, row_idx, col_idx, max_col)
                if value and parse_num(value) is None:
                    attrs["name"] = str(value).strip()
                    found_name = True
                continue

            if not found_profession and key in ("职业", "本职", "职业名称", "occupation"):
                value = get_next_value(target_sheet, row_idx, col_idx, max_col)
                if value and parse_num(value) is None:
                    attrs["profession"] = str(value).strip()
                    found_profession = True
                continue

            if key in ("san", "理智", "理智值"):
                value = get_next_value(target_sheet, row_idx, col_idx, max_col, expect_num=True)
                if value is not None:
                    attrs["san"] = max(attrs["san"], int(value))
                continue

            if key in ("幸运", "luck"):
                value = get_next_value(target_sheet, row_idx, col_idx, max_col, expect_num=True)
                if value is not None:
                    attrs["luck"] = max(attrs["luck"], int(value))
                continue

            for base_name, aliases in base_keys.items():
                if any(alias in key for alias in aliases) and len(key) <= 10:
                    value = get_base_value(target_sheet, row_idx, col_idx, max_col)
                    if value:
                        attrs["bases"][base_name] = max(attrs["bases"][base_name], value)

            if 1 <= len(key) <= 18 and any(alias in key for alias in all_skill_aliases):
                skill_value = get_skill_total(target_sheet, row_idx, col_idx, max_col)
                if skill_value is not None:
                    skill_key = normalize_skill_key(key)
                    attrs["skills"][skill_key] = max(attrs["skills"].get(skill_key, 0), skill_value)

    workbook.close()
    if repaired_path and os.path.exists(repaired_path):
        os.remove(repaired_path)

    for base_name, value in attrs["bases"].items():
        if value <= 0:
            attrs["bases"][base_name] = 50
    if attrs["san"] <= 0:
        attrs["san"] = attrs["bases"]["pow"]
    if attrs["luck"] <= 0:
        attrs["luck"] = attrs["bases"]["pow"]
    attrs["name"] = truncate_name(attrs["name"], 12)
    attrs["profession"] = truncate_name(attrs["profession"], 10)
    return attrs


def truncate_name(value: str, max_len: int = 12) -> str:
    value = str(value or "").strip()
    if len(value) <= max_len:
        return value
    for sep in ("·", "・", "•", "-", "‐", "—"):
        if sep in value:
            parts = [part for part in value.split(sep) if part]
            result = parts[0]
            for part in parts[1:]:
                candidate = f"{result}{sep}{part}"
                if len(candidate) <= max_len:
                    result = candidate
            return result
    return value[: max_len - 1] + "…"


def default_for_alias(alias: str, bases: dict[str, int]) -> int:
    if "闪避" in alias:
        return bases.get("dex", 50) // 2
    if "母语" in alias:
        return bases.get("edu", 50)
    for key, default in SKILL_DEFAULTS.items():
        if key in alias or alias in key:
            return default
    return 1


def skill_value(skills: dict[str, int], aliases: list[str], bases: dict[str, int]) -> int:
    best = 0
    for alias in aliases:
        alias_key = clean_key(alias)
        for skill_name, value in skills.items():
            if alias_key in skill_name:
                best = max(best, int(value))
        best = max(best, default_for_alias(alias, bases))
    return min(best, 99)


def skill_bonus(skills: dict[str, int], aliases: list[str], bases: dict[str, int], cap: int = 145) -> int:
    investments = []
    for alias in aliases:
        alias_key = clean_key(alias)
        for skill_name, value in skills.items():
            if alias_key in skill_name:
                default = default_for_alias(alias, bases)
                investments.append(max(0, int(value) - default))
    if not investments:
        return 0
    investments.sort(reverse=True)
    primary = investments[0] * 1.55
    support = math.sqrt(sum(investments[1:])) * 7.0 if len(investments) > 1 else 0
    return int(min(cap, primary + support))


def clamp_stat(value: float) -> int:
    return max(0, min(1200, int(round(value))))


def stat_to_rank(value: int) -> str:
    if value >= 1100:
        return "SS"
    if value >= 900:
        return "S"
    if value >= 750:
        return "A"
    if value >= 600:
        return "B"
    if value >= 450:
        return "C"
    if value >= 300:
        return "D"
    return "E"


def total_rank(points: int) -> str:
    for rank, threshold in RANK_THRESHOLDS:
        if points >= threshold:
            return rank
    return "G"


def average_top(values: list[int], count: int) -> float:
    values = sorted(values, reverse=True)
    return sum(values[:count]) / max(1, min(count, len(values)))


def aptitude_grade(score: float) -> str:
    if score >= 90:
        return "S"
    if score >= 75:
        return "A"
    if score >= 60:
        return "B"
    if score >= 45:
        return "C"
    if score >= 30:
        return "D"
    if score >= 15:
        return "E"
    if score >= 5:
        return "F"
    return "G"


def calculate_aptitudes(attrs: dict[str, Any]) -> dict[str, dict[str, Any]]:
    bases = attrs["bases"]
    skills = attrs["skills"]

    def best(key: str) -> int:
        return skill_value(skills, APTITUDE_GROUPS[key], bases)

    def values(key: str) -> list[int]:
        return [skill_value(skills, [alias], bases) for alias in APTITUDE_GROUPS[key]]

    mythos = skill_value(skills, ["克苏鲁"], bases)
    scores = {
        "melee": best("melee") * 0.55 + bases["str"] * 0.25 + bases["dex"] * 0.20,
        "ranged": best("ranged") * 0.65 + bases["dex"] * 0.35,
        "city": average_top(values("city"), 3) * 0.75 + bases["int"] * 0.15 + bases["app"] * 0.10,
        "wild": average_top(values("wild"), 3) * 0.75 + bases["con"] * 0.15 + bases["dex"] * 0.10,
        "library": average_top(values("library"), 3) * 0.75 + bases["edu"] * 0.25,
        "data": average_top(values("data"), 3) * 0.80 + bases["edu"] * 0.20,
        "endurance": bases["con"] * 0.60 + average_top(values("endurance"), 2) * 0.40,
        "emergency": average_top(values("emergency"), 3) * 0.75 + bases["dex"] * 0.15 + bases["int"] * 0.10,
        "mental": bases["pow"] * 0.45 + attrs["san"] * 0.35 + mythos * 0.20,
        "luck": attrs["luck"] * 0.70 + bases["pow"] * 0.30,
    }
    return {key: {"score": round(min(100, max(0, score)), 1), "grade": aptitude_grade(score)} for key, score in scores.items()}


def calculate_uma(attrs: dict[str, Any]) -> dict[str, Any]:
    bases = attrs["bases"]
    skills = attrs["skills"]

    stats = {
        "speed": clamp_stat(((bases["dex"] * 2 + bases["str"]) / 3) * 12 + skill_bonus(skills, STAT_SKILL_GROUPS["speed"], bases)),
        "stamina": clamp_stat(((bases["con"] * 2 + bases["str"] + bases["siz"]) / 4) * 12 + skill_bonus(skills, STAT_SKILL_GROUPS["stamina"], bases)),
        "power": clamp_stat(((bases["str"] * 2 + bases["con"]) / 3) * 12 + skill_bonus(skills, STAT_SKILL_GROUPS["power"], bases)),
        "guts": clamp_stat(((bases["pow"] * 2 + bases["con"]) / 3) * 12 + skill_bonus(skills, STAT_SKILL_GROUPS["guts"], bases)),
        "wisdom": clamp_stat(((bases["int"] + bases["edu"] + bases["app"]) / 3) * 12 + skill_bonus(skills, STAT_SKILL_GROUPS["wisdom"], bases)),
    }
    stat_ranks = {key: stat_to_rank(value) for key, value in stats.items()}
    aptitudes = calculate_aptitudes(attrs)
    apt_avg = sum(item["score"] for item in aptitudes.values()) / len(aptitudes)
    stat_values = sorted(stats.values())
    eval_points = int(sum(stat_values) * 3.4 + sum(stat_values[-2:]) * 1.5 + apt_avg * 18)
    rank = total_rank(eval_points)

    return {
        "stats": stats,
        "stat_ranks": stat_ranks,
        "aptitudes": aptitudes,
        "eval_points": eval_points,
        "rank": rank,
    }


def draw_text_fit(
    draw: ImageDraw.ImageDraw,
    box: list[int],
    text: str,
    max_size: int,
    *,
    fill=BROWN,
    anchor="mm",
    stroke_width=0,
    stroke_fill=WHITE,
    min_size=15,
) -> None:
    x0, y0, x1, y1 = box
    max_w = x1 - x0
    max_h = y1 - y0
    size = max_size
    while size >= min_size:
        current_font = get_font(size)
        bbox = draw.textbbox((0, 0), text, font=current_font, stroke_width=stroke_width)
        if bbox[2] - bbox[0] <= max_w and bbox[3] - bbox[1] <= max_h:
            break
        size -= 2
    pos = (x0, (y0 + y1) / 2) if anchor == "lm" else ((x0 + x1) / 2, (y0 + y1) / 2)
    draw.text(pos, text, font=get_font(size), fill=fill, anchor=anchor, stroke_width=stroke_width, stroke_fill=stroke_fill)


def paste_fit(base: Image.Image, overlay_path: Path, box: list[int]) -> None:
    if not overlay_path.exists():
        raise FileNotFoundError(f"缺少素材: {overlay_path}")
    overlay = Image.open(overlay_path).convert("RGBA")
    width = box[2] - box[0]
    height = box[3] - box[1]
    overlay = overlay.resize((width, height), Image.Resampling.LANCZOS)
    base.alpha_composite(overlay, (box[0], box[1]))


def render_uma_card(attrs: dict[str, Any], result: dict[str, Any], output_path: str | Path) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"缺少模板: {TEMPLATE_PATH}")
    image = Image.open(TEMPLATE_PATH).convert("RGBA")
    draw = ImageDraw.Draw(image)

    paste_fit(image, RANK_DIR / f"rank_{result['rank']}.png", FIELDS["rank_badge"]["box"])
    draw_text_fit(draw, FIELDS["character_name"]["box"], attrs["name"], FIELDS["character_name"]["font"], stroke_width=3)
    if "player_name" in FIELDS:
        player_text = f"玩家：{attrs.get('player') or '-'}"
        draw_text_fit(draw, FIELDS["player_name"]["box"], player_text, FIELDS["player_name"]["font"], stroke_width=1)
    draw_text_fit(
        draw,
        FIELDS["profession"]["box"],
        attrs["profession"],
        FIELDS["profession"]["font"],
        fill=WHITE,
        stroke_width=2,
        stroke_fill=(176, 92, 0),
    )
    draw_text_fit(draw, FIELDS["eval_points"]["box"], f"{result['eval_points']:,}", FIELDS["eval_points"]["font"], anchor="lm")

    for stat_field in FIELDS["stats"]:
        key = stat_field["key"]
        rank = result["stat_ranks"][key]
        value = result["stats"][key]
        paste_fit(image, STAT_DIR / f"stat_{rank}.png", stat_field["badge_box"])
        draw_text_fit(draw, stat_field["value_box"], str(value), 43)

    for item in FIELDS["aptitudes"]:
        grade = result["aptitudes"][item["key"]]["grade"]
        box = item["box"]
        label_box = [box[0] + 18, box[1] + 14, item["grade_box"][0] - 8, box[3] - 14]
        draw_text_fit(draw, label_box, item["label"], 27, anchor="lm")
        draw_text_fit(
            draw,
            item["grade_box"],
            grade,
            42 if grade != "SS" else 34,
            fill=GRADE_COLORS.get(grade, BROWN),
            stroke_width=2,
            stroke_fill=(255, 239, 224),
        )

    output_path = Path(output_path)
    output_path.parent.mkdir(exist_ok=True)
    image.convert("RGB").save(output_path)
    return output_path


def sanitize_filename(value: str) -> str:
    safe = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", str(value)).strip().rstrip(".")
    return safe or "调查员"


def safe_path_key(value: str, fallback: str = "unknown") -> str:
    safe = sanitize_filename(value)
    safe = re.sub(r"\s+", "_", safe)
    return safe or fallback


def read_json_file(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return default


def write_json_file(path: Path, data: Any) -> None:
    path.parent.mkdir(exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def stable_user_dir(user_key: str) -> Path:
    path = STABLE_DIR / safe_path_key(user_key)
    path.mkdir(parents=True, exist_ok=True)
    return path


def stable_card_path(user_key: str, card_id: str) -> Path:
    return stable_user_dir(user_key) / f"{safe_path_key(card_id)}.json"


def card_summary(card: dict[str, Any]) -> dict[str, Any]:
    return {
        "card_id": card.get("card_id", ""),
        "name": card.get("name", ""),
        "profession": card.get("profession", ""),
        "rank": card.get("rank", ""),
        "eval_points": card.get("eval_points", 0),
        "stats": card.get("stats", {}),
        "stat_ranks": card.get("stat_ranks", {}),
        "aptitudes": card.get("aptitudes", {}),
        "image_name": card.get("image_name", ""),
        "created_at": card.get("created_at", ""),
    }


def save_stable_card(
    user_key: str,
    attrs: dict[str, Any],
    result: dict[str, Any],
    image_name: str,
    group_id: str = "",
    user_name: str = "",
) -> dict[str, Any] | None:
    if not user_key:
        return None

    card = {
        "card_id": image_name,
        "name": attrs["name"],
        "profession": attrs["profession"],
        "player": attrs.get("player", ""),
        "rank": result["rank"],
        "eval_points": result["eval_points"],
        "stats": result["stats"],
        "stat_ranks": result["stat_ranks"],
        "aptitudes": {key: value["grade"] for key, value in result["aptitudes"].items()},
        "image_name": image_name,
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
    write_json_file(stable_card_path(user_key, image_name), card)
    if group_id:
        register_group_user(group_id, user_key, user_name)
    return card


def load_stable_cards(user_key: str) -> list[dict[str, Any]]:
    path = stable_user_dir(user_key)
    cards = []
    for item in path.glob("*.json"):
        data = read_json_file(item, None)
        if isinstance(data, dict) and data.get("card_id"):
            cards.append(data)
    cards.sort(key=lambda card: (card.get("created_at", ""), card.get("card_id", "")), reverse=True)
    return cards


def find_stable_card(user_key: str, query: str) -> dict[str, Any] | None:
    query = str(query or "").strip()
    if not query:
        return None

    cards = load_stable_cards(user_key)
    for card in cards:
        if query == card.get("card_id") or query == card.get("name"):
            return card

    matches = [
        card for card in cards
        if query in str(card.get("name", "")) or query in str(card.get("card_id", ""))
    ]
    if len(matches) == 1:
        return matches[0]
    if len(matches) > 1:
        return {"__ambiguous__": [card.get("name", card.get("card_id", "")) for card in matches]}
    return None


def remove_stable_card(user_key: str, query: str) -> tuple[dict[str, Any] | None, str]:
    card = find_stable_card(user_key, query)
    if not card:
        return None, "没有找到这张马卡。"
    if "__ambiguous__" in card:
        names = "、".join(card["__ambiguous__"][:6])
        return None, f"找到多张同名马卡：{names}。请换更完整的名字。"

    path = stable_card_path(user_key, card["card_id"])
    if path.exists():
        path.unlink()

    team = load_horse_team(user_key)
    changed = False
    for slot, card_id in list(team.items()):
        if card_id == card["card_id"]:
            team[slot] = ""
            changed = True
    if changed:
        save_horse_team(user_key, team)
    return card, f"{card['name']} 已从马房删除。"


def horse_team_path(user_key: str) -> Path:
    return HORSE_TEAM_DIR / f"{safe_path_key(user_key)}.json"


def load_horse_team(user_key: str) -> dict[str, str]:
    data = read_json_file(horse_team_path(user_key), {})
    team = {slot: str(data.get(slot) or "") for slot in TEAM_SLOTS}
    return team


def save_horse_team(user_key: str, team: dict[str, str]) -> None:
    payload = {slot: team.get(slot, "") for slot in TEAM_SLOTS}
    payload["updated_at"] = dt.datetime.now().isoformat(timespec="seconds")
    write_json_file(horse_team_path(user_key), payload)


def slot_from_text(value: str) -> str | None:
    return SLOT_ALIASES.get(str(value or "").strip().lower()) or SLOT_ALIASES.get(str(value or "").strip())


def parse_order(value: str) -> list[str] | None:
    raw = str(value or "").strip()
    if not raw:
        return None

    compact = re.sub(r"[\s,，、/|]+", "", raw)
    if len(compact) == 3 and all(char in ORDER_CHAR_MAP for char in compact):
        order = [ORDER_CHAR_MAP[char] for char in compact]
        return order if len(set(order)) == 3 else None

    tokens = [token for token in re.split(r"[\s,，、/|]+", raw) if token]
    order = [slot_from_text(token) for token in tokens]
    if len(order) == 3 and all(order) and len(set(order)) == 3:
        return [str(slot) for slot in order]
    return None


def team_detail(user_key: str) -> dict[str, Any]:
    team = load_horse_team(user_key)
    cards = {card["card_id"]: card for card in load_stable_cards(user_key)}
    slots = {}
    for slot, label in TEAM_SLOTS.items():
        card_id = team.get(slot, "")
        slots[slot] = {
            "label": label,
            "card_id": card_id,
            "card": card_summary(cards[card_id]) if card_id in cards else None,
        }
    return {"slots": slots, "complete": all(item["card"] for item in slots.values())}


def complete_team_cards(user_key: str) -> dict[str, dict[str, Any]] | None:
    detail = team_detail(user_key)
    if not detail["complete"]:
        return None
    return {slot: detail["slots"][slot]["card"] for slot in TEAM_SLOTS}


def group_path(group_id: str) -> Path:
    return RACE_GROUP_DIR / f"{safe_path_key(group_id)}.json"


def register_group_user(group_id: str, user_key: str, user_name: str = "") -> None:
    if not group_id or not user_key:
        return
    data = read_json_file(group_path(group_id), {"group_id": group_id, "users": {}})
    users = data.setdefault("users", {})
    users[user_key] = {
        "name": user_name or users.get(user_key, {}).get("name") or f"玩家{user_key}",
        "updated_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
    data["group_id"] = group_id
    write_json_file(group_path(group_id), data)


def choose_opponent(group_id: str, user_key: str, rng: random.Random) -> tuple[str, str] | None:
    data = read_json_file(group_path(group_id), {"users": {}})
    candidates = []
    own_power = team_power(user_key)
    for candidate_key, info in data.get("users", {}).items():
        if candidate_key == user_key:
            continue
        if complete_team_cards(candidate_key):
            gap = abs(team_power(candidate_key) - own_power)
            candidates.append((gap, rng.random(), candidate_key, info.get("name") or f"玩家{candidate_key}"))
    if not candidates:
        return None
    candidates.sort()
    return candidates[0][2], candidates[0][3]


def team_power(user_key: str) -> int:
    cards = complete_team_cards(user_key)
    if not cards:
        return 0
    return sum(int(card.get("eval_points", 0)) for card in cards.values())


def performance_score(card: dict[str, Any], rng: random.Random) -> float:
    stats = card.get("stats", {})
    speed = int(stats.get("speed", 0))
    stamina = int(stats.get("stamina", 0))
    power = int(stats.get("power", 0))
    guts = int(stats.get("guts", 0))
    wisdom = int(stats.get("wisdom", 0))
    base = speed * 0.34 + stamina * 0.20 + power * 0.22 + guts * 0.14 + wisdom * 0.10
    spread = max(12.0, 170.0 - wisdom * 0.09 - stamina * 0.04)
    score = base + rng.uniform(-spread, spread)
    if score < base:
        score += min(base - score, guts * 0.08)
    return round(score, 1)


def run_race(
    user_key: str,
    group_id: str,
    order_text: str,
    *,
    user_name: str = "",
    opponent_key: str = "",
    opponent_name: str = "",
    seed: str = "",
) -> dict[str, Any]:
    order = parse_order(order_text)
    if not order:
        return {"status": "error", "msg": "出场顺序需要包含上、中、下三项，例如 上中下 或 下 上 中"}

    own_cards = complete_team_cards(user_key)
    if not own_cards:
        return {"status": "error", "msg": "你的马队还没有设置完整。请先设置上等马、中等马、下等马。"}

    rng = random.Random(seed) if seed else random.Random()
    if group_id:
        register_group_user(group_id, user_key, user_name)

    if opponent_key:
        if opponent_key == user_key:
            return {"status": "error", "msg": "不能挑战自己。"}
        opponent_cards = complete_team_cards(opponent_key)
        if not opponent_cards:
            return {"status": "error", "msg": "对手马队还没有设置完整。"}
        opponent_label = opponent_name or f"玩家{opponent_key}"
    else:
        picked = choose_opponent(group_id, user_key, rng)
        if not picked:
            return {"status": "error", "msg": "当前群没有可自动匹配的完整马队。"}
        opponent_key, opponent_label = picked
        opponent_cards = complete_team_cards(opponent_key)

    if not opponent_cards:
        return {"status": "error", "msg": "对手马队还没有设置完整。"}

    opponent_order = list(TEAM_SLOTS)
    rng.shuffle(opponent_order)

    own_wins = 0
    opponent_wins = 0
    rounds = []
    round_names = ("第一场", "第二场", "第三场")
    for index, (own_slot, opponent_slot) in enumerate(zip(order, opponent_order)):
        own_card = own_cards[own_slot]
        opponent_card = opponent_cards[opponent_slot]
        own_score = performance_score(own_card, rng)
        opponent_score = performance_score(opponent_card, rng)
        if round(own_score) > round(opponent_score):
            winner = "user"
        elif round(opponent_score) > round(own_score):
            winner = "opponent"
        elif int(own_card.get("eval_points", 0)) > int(opponent_card.get("eval_points", 0)):
            winner = "user"
        elif int(opponent_card.get("eval_points", 0)) > int(own_card.get("eval_points", 0)):
            winner = "opponent"
        else:
            winner = "user" if rng.random() >= 0.5 else "opponent"

        if winner == "user":
            own_wins += 1
            winner_name = own_card["name"]
        else:
            opponent_wins += 1
            winner_name = opponent_card["name"]

        rounds.append({
            "round": index + 1,
            "round_name": round_names[index],
            "user_slot": own_slot,
            "opponent_slot": opponent_slot,
            "user_horse": own_card,
            "opponent_horse": opponent_card,
            "user_score": own_score,
            "opponent_score": opponent_score,
            "winner": winner,
            "winner_name": winner_name,
            "text": f"{round_names[index]}：{own_card['name']} 对 {opponent_card['name']}。{winner_name} 先过线。",
        })

    winner_side = "user" if own_wins > opponent_wins else "opponent"
    race_id = f"race_{safe_path_key(user_key)}_{safe_path_key(opponent_key)}_{dt.datetime.now().strftime('%Y%m%d_%H%M%S_%f')}"
    race = {
        "id": race_id,
        "user_key": user_key,
        "user_name": user_name or f"玩家{user_key}",
        "opponent_key": opponent_key,
        "opponent_name": opponent_label,
        "order": order,
        "opponent_order": opponent_order,
        "score": f"{own_wins}:{opponent_wins}",
        "winner": "",
        "winner_side": winner_side,
        "rounds": rounds,
        "created_at": dt.datetime.now().isoformat(timespec="seconds"),
    }
    race["winner"] = race["user_name"] if winner_side == "user" else race["opponent_name"]
    race["img"] = render_race_result(race)
    write_json_file(RACE_DIR / f"{race_id}.json", race)
    return {
        "status": "ok",
        "msg": f"{race['winner']}赢下马赛。",
        "race_id": race_id,
        "img": race["img"],
        "score": race["score"],
        "winner": race["winner"],
        "winner_side": race["winner_side"],
        "opponent_key": opponent_key,
        "opponent_name": race["opponent_name"],
        "rounds": [
            {
                "round": item["round"],
                "text": item["text"],
                "user_score": item["user_score"],
                "opponent_score": item["opponent_score"],
                "winner": item["winner"],
            }
            for item in rounds
        ],
    }


def render_race_result(race: dict[str, Any]) -> str:
    width, height = 1180, 760
    image = Image.new("RGB", (width, height), (249, 244, 235))
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, width, 118), fill=(132, 44, 38))
    draw.rectangle((0, 118, width, 128), fill=(216, 161, 64))
    draw.text((48, 34), "马赛结果", font=get_font(48), fill=WHITE)
    draw.text((48, 92), f"{race['user_name']}  {race['score']}  {race['opponent_name']}", font=get_font(26), fill=(255, 242, 210))
    draw.text((900, 45), f"胜者：{race['winner']}", font=get_font(32), fill=WHITE)

    y = 164
    for item in race["rounds"]:
        fill = (255, 255, 255) if item["round"] % 2 else (244, 237, 224)
        draw.rounded_rectangle((44, y, width - 44, y + 142), radius=18, fill=fill, outline=(218, 197, 164), width=3)
        draw.text((72, y + 22), item["round_name"], font=get_font(31), fill=(103, 45, 0))
        draw.text(
            (72, y + 76),
            f"{TEAM_SLOTS[item['user_slot']]}：{item['user_horse']['name']}  {item['user_score']}",
            font=get_font(27),
            fill=(46, 82, 126),
        )
        draw.text(
            (520, y + 76),
            f"{TEAM_SLOTS[item['opponent_slot']]}：{item['opponent_horse']['name']}  {item['opponent_score']}",
            font=get_font(27),
            fill=(122, 57, 49),
        )
        draw.text((920, y + 74), f"{item['winner_name']}胜", font=get_font(30), fill=(146, 88, 12))
        y += 168

    image_name = race["id"]
    output = SAVE_DIR / f"{image_name}.png"
    image.save(output)
    return image_name


def build_card_from_file(
    file_path: str,
    *,
    user_key: str = "",
    group_id: str = "",
    user_name: str = "",
) -> dict[str, Any]:
    attrs = parse_coc_excel(file_path)
    result = calculate_uma(attrs)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    title = f"{sanitize_filename(attrs['name'])}_uma_{timestamp}"
    output = SAVE_DIR / f"{title}.png"
    render_uma_card(attrs, result, output)
    stable_card = save_stable_card(user_key, attrs, result, title, group_id, user_name)
    return {
        "status": "ok",
        "name": title,
        "card_id": stable_card["card_id"] if stable_card else "",
        "stable_saved": bool(stable_card),
        "character": attrs["name"],
        "profession": attrs["profession"],
        "eval_points": result["eval_points"],
        "rank": result["rank"],
        "stats": result["stats"],
        "stat_ranks": result["stat_ranks"],
        "aptitudes": {key: value["grade"] for key, value in result["aptitudes"].items()},
    }


def download_excel_to_temp(url: str, suffix: str = ".xlsx") -> str:
    fd, temp_path = tempfile.mkstemp(prefix="uma_download_", suffix=suffix)
    os.close(fd)
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    Path(temp_path).write_bytes(response.content)
    return temp_path


def excel_suffix_from_name(value: str) -> str:
    lower = str(value or "").lower()
    if lower.endswith(".xls") and not lower.endswith(".xlsx"):
        return ".xls"
    return ".xlsx"


def request_payload() -> dict[str, Any]:
    payload = request.get_json(silent=True) or {}
    if request.form:
        payload.update(request.form.to_dict())
    payload.update(request.args.to_dict())
    return payload


@app.route("/stable/list", methods=["GET", "POST"])
def stable_list():
    payload = request_payload()
    user_key = str(payload.get("user_key") or "").strip()
    if not user_key:
        return jsonify({"status": "error", "msg": "缺少 user_key"})
    cards = [card_summary(card) for card in load_stable_cards(user_key)]
    return jsonify({"status": "ok", "cards": cards, "count": len(cards)})


@app.route("/stable/detail", methods=["GET", "POST"])
def stable_detail():
    payload = request_payload()
    user_key = str(payload.get("user_key") or "").strip()
    query = str(payload.get("name") or payload.get("card_id") or "").strip()
    if not user_key or not query:
        return jsonify({"status": "error", "msg": "缺少 user_key 或 name"})
    card = find_stable_card(user_key, query)
    if not card:
        return jsonify({"status": "error", "msg": "没有找到这张马卡。"})
    if "__ambiguous__" in card:
        names = "、".join(card["__ambiguous__"][:6])
        return jsonify({"status": "error", "msg": f"找到多张同名马卡：{names}。请换更完整的名字。"})
    return jsonify({"status": "ok", "card": card_summary(card)})


@app.route("/stable/remove", methods=["GET", "POST"])
def stable_remove():
    payload = request_payload()
    user_key = str(payload.get("user_key") or "").strip()
    query = str(payload.get("name") or payload.get("card_id") or "").strip()
    if not user_key or not query:
        return jsonify({"status": "error", "msg": "缺少 user_key 或 name"})
    card, msg = remove_stable_card(user_key, query)
    if not card:
        return jsonify({"status": "error", "msg": msg})
    return jsonify({"status": "ok", "msg": msg, "card": card_summary(card), "team": team_detail(user_key)})


@app.route("/horse_team/set", methods=["GET", "POST"])
def horse_team_set():
    payload = request_payload()
    user_key = str(payload.get("user_key") or "").strip()
    slot = slot_from_text(str(payload.get("slot") or ""))
    query = str(payload.get("name") or payload.get("card_id") or "").strip()
    group_id = str(payload.get("group_id") or "").strip()
    user_name = str(payload.get("user_name") or "").strip()
    if not user_key or not slot or not query:
        return jsonify({"status": "error", "msg": "缺少 user_key、slot 或 name"})
    card = find_stable_card(user_key, query)
    if not card:
        return jsonify({"status": "error", "msg": "没有找到这张马卡。"})
    if "__ambiguous__" in card:
        names = "、".join(card["__ambiguous__"][:6])
        return jsonify({"status": "error", "msg": f"找到多张同名马卡：{names}。请换更完整的名字。"})
    team = load_horse_team(user_key)
    team[slot] = card["card_id"]
    save_horse_team(user_key, team)
    if group_id:
        register_group_user(group_id, user_key, user_name)
    return jsonify({
        "status": "ok",
        "msg": f"{card['name']} 已设为{TEAM_SLOTS[slot]}马。",
        "team": team_detail(user_key),
    })


@app.route("/horse_team/list", methods=["GET", "POST"])
def horse_team_list():
    payload = request_payload()
    user_key = str(payload.get("user_key") or "").strip()
    group_id = str(payload.get("group_id") or "").strip()
    user_name = str(payload.get("user_name") or "").strip()
    if not user_key:
        return jsonify({"status": "error", "msg": "缺少 user_key"})
    if group_id:
        register_group_user(group_id, user_key, user_name)
    return jsonify({"status": "ok", "team": team_detail(user_key)})


@app.route("/race/run", methods=["GET", "POST"])
def race_run():
    payload = request_payload()
    user_key = str(payload.get("user_key") or "").strip()
    group_id = str(payload.get("group_id") or "").strip()
    order = str(payload.get("order") or "").strip()
    if not user_key or not group_id or not order:
        return jsonify({"status": "error", "msg": "缺少 user_key、group_id 或 order"})
    result = run_race(
        user_key,
        group_id,
        order,
        user_name=str(payload.get("user_name") or "").strip(),
        opponent_key=str(payload.get("opponent_key") or "").strip(),
        opponent_name=str(payload.get("opponent_name") or "").strip(),
        seed=str(payload.get("seed") or "").strip(),
    )
    return jsonify(result)


@app.route("/generate_uma", methods=["GET", "POST"])
def generate_uma():
    temp_path = None
    try:
        payload = request_payload()
        if request.method == "POST" and "file" in request.files:
            uploaded = request.files["file"]
            suffix = Path(uploaded.filename or "card.xlsx").suffix or ".xlsx"
            fd, temp_path = tempfile.mkstemp(prefix="uma_upload_", suffix=suffix)
            os.close(fd)
            uploaded.save(temp_path)
        else:
            url = request.args.get("url", "").strip()
            if not url:
                return jsonify({"status": "error", "msg": "缺少 url 或上传文件"})
            temp_path = download_excel_to_temp(url, excel_suffix_from_name(url))

        return jsonify(build_card_from_file(
            temp_path,
            user_key=str(payload.get("user_key") or "").strip(),
            group_id=str(payload.get("group_id") or "").strip(),
            user_name=str(payload.get("user_name") or "").strip(),
        ))
    except Exception as exc:
        return jsonify({"status": "error", "msg": str(exc)})
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)


@app.route("/generate_uma_group_file", methods=["GET", "POST"])
def generate_uma_group_file():
    temp_path = None
    try:
        payload = request.get_json(silent=True) or {}
        api_base = (payload.get("api_base") or request.args.get("api_base") or "http://luckylillia:3010/api").strip().rstrip("/")
        group_id = payload.get("group_id") or request.args.get("group_id")
        file_id = payload.get("file_id") or request.args.get("file_id")
        filename = payload.get("filename") or request.args.get("filename") or "card.xlsx"
        if not group_id or not file_id:
            return jsonify({"status": "error", "msg": "缺少 group_id 或 file_id"})

        api_resp = requests.post(
            f"{api_base}/get_group_file_download_url",
            json={"group_id": int(group_id), "file_id": str(file_id)},
            timeout=20,
        )
        api_resp.raise_for_status()
        api_json = api_resp.json()
        download_url = (api_json.get("data") or {}).get("download_url")
        if not download_url:
            return jsonify({"status": "error", "msg": f"获取文件下载链接失败：{api_json.get('message') or api_json}"})

        temp_path = download_excel_to_temp(download_url, excel_suffix_from_name(filename))
        return jsonify(build_card_from_file(
            temp_path,
            user_key=str(payload.get("user_key") or request.args.get("user_key") or "").strip(),
            group_id=str(payload.get("race_group_id") or payload.get("group_id") or request.args.get("race_group_id") or group_id or "").strip(),
            user_name=str(payload.get("user_name") or request.args.get("user_name") or "").strip(),
        ))
    except Exception as exc:
        return jsonify({"status": "error", "msg": str(exc)})
    finally:
        if temp_path and os.path.exists(temp_path):
            os.remove(temp_path)


@app.route("/get_img")
def get_img():
    title = request.args.get("title", "")
    path = SAVE_DIR / f"{title}.png"
    if path.exists():
        return send_file(path, mimetype="image/png")
    return "Not Found", 404


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=21999, threaded=False)
