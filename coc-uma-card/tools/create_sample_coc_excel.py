from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "samples"
OUT_FILES = [
    OUT_DIR / "sample_coc_character.xlsx",
    OUT_DIR / "sample_coc_character_vertical.xlsx",
    OUT_DIR / "sample_coc_character_compact.xlsx",
]


def save_first_sample() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "调查员"

    rows = [
        ["姓名", "雾岛莲", "玩家", "玩家A", "职业", "私家侦探"],
        ["STR", 65, "CON", 70, "POW", 80],
        ["DEX", 75, "INT", 85, "SIZ", 60],
        ["APP", 65, "EDU", 80, "SAN", 76],
        ["幸运", 70],
        [],
        ["技能", "成功率", "困难", "极难"],
        ["侦查", 78, 39, 15],
        ["聆听", 65, 32, 13],
        ["心理学", 70, 35, 14],
        ["图书馆使用", 82, 41, 16],
        ["历史", 60, 30, 12],
        ["神秘学", 55, 27, 11],
        ["计算机使用", 45, 22, 9],
        ["会计", 35, 17, 7],
        ["法律", 48, 24, 9],
        ["斗殴", 68, 34, 13],
        ["手枪", 42, 21, 8],
        ["投掷", 55, 27, 11],
        ["闪避", 62, 31, 12],
        ["急救", 58, 29, 11],
        ["医学", 35, 17, 7],
        ["机械维修", 40, 20, 8],
        ["电气维修", 25, 12, 5],
        ["生存", 50, 25, 10],
        ["追踪", 48, 24, 9],
        ["导航", 45, 22, 9],
        ["克苏鲁神话", 8, 4, 1],
    ]

    for row in rows:
        ws.append(row)

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 16

    wb.save(OUT_FILES[0])


def save_second_sample() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "人物核心"
    rows = [
        ["调查员", "白石遥"],
        ["玩家名", "玩家B"],
        ["职业名称", "医生"],
        ["力量", 50],
        ["体质", 80],
        ["意志", 75],
        ["敏捷", 55],
        ["智力", 70],
        ["体型", 60],
        ["外貌", 60],
        ["教育", 90],
        ["理智", 72],
        ["幸运", 65],
        [],
        ["急救", 85, 42, 17],
        ["医学", 78, 39, 15],
        ["精神分析", 65, 32, 13],
        ["图书馆", 70, 35, 14],
        ["侦查", 60, 30, 12],
        ["聆听", 55, 27, 11],
        ["心理学", 75, 37, 15],
        ["会计", 20, 10, 4],
        ["法律", 28, 14, 5],
        ["手枪", 30, 15, 6],
        ["斗殴", 45, 22, 9],
        ["闪避", 42, 21, 8],
        ["生存", 35, 17, 7],
    ]
    for row in rows:
        ws.append(row)
    for col in "ABCD":
        ws.column_dimensions[col].width = 16
    wb.save(OUT_FILES[1])


def save_third_sample() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "角色卡"
    rows = [
        ["姓名", "", "黑泽真"],
        ["PL", "", "玩家C"],
        ["职业", "", "记者"],
        ["STR", "", 45, "DEX", "", 85],
        ["CON", "", 55, "SIZ", "", 50],
        ["POW", "", 60, "INT", "", 90],
        ["APP", "", 75, "EDU", "", 85],
        ["SAN", "", 58, "Luck", "", 80],
        [],
        ["技能名", "初始", "成长", "成功率", "困难", "极难"],
        ["话术", 5, 55, 60, 30, 12],
        ["说服", 10, 45, 55, 27, 11],
        ["取悦", 15, 35, 50, 25, 10],
        ["侦查", 25, 45, 70, 35, 14],
        ["聆听", 20, 40, 60, 30, 12],
        ["图书馆使用", 20, 60, 80, 40, 16],
        ["历史", 5, 55, 60, 30, 12],
        ["外语", 1, 49, 50, 25, 10],
        ["法律", 5, 40, 45, 22, 9],
        ["潜行", 20, 45, 65, 32, 13],
        ["追踪", 10, 35, 45, 22, 9],
        ["射击", 20, 50, 70, 35, 14],
        ["投掷", 20, 35, 55, 27, 11],
        ["克苏鲁神话", 0, 12, 12, 6, 2],
    ]
    for row in rows:
        ws.append(row)
    for col in "ABCDEF":
        ws.column_dimensions[col].width = 14
    wb.save(OUT_FILES[2])


def main() -> None:
    OUT_DIR.mkdir(exist_ok=True)
    save_first_sample()
    save_second_sample()
    save_third_sample()
    for path in OUT_FILES:
        print(path)


if __name__ == "__main__":
    main()
